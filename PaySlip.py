import xlsxwriter
import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import importlib
import PDF as Fi
import Email as Em
import time

wb = load_workbook(r'PAYROLL.xlsx')
ws = wb.active
wg = wb['Pay Slip']
Sheet_row = ws.max_row + 1
r = 2
while r < Sheet_row:
    print("Start Cell")
    wg['D12'] = ws.cell(row = r,column = 1).value
    wg['D13'] = ws.cell(row = r,column = 2).value
    wg['D14'] = ws.cell(row = r,column = 3).value
    wg['D15'] = ws.cell(row = r,column = 4).value
    wg['D16'] = ws.cell(row = r,column = 5).value
    wg['F12'] = ws.cell(row = r,column = 6).value
    wg['F13'] = ws.cell(row = r,column = 7).value
    wg['F14'] = ws.cell(row = r,column = 8).value
    wg['F15'] = ws.cell(row = r,column = 9).value
    wg['F16'] = ws.cell(row = r,column = 10).value
    wg['D19'] = ws.cell(row = r,column = 11).value
    wg['D20'] = ws.cell(row = r,column = 12).value
    wg['D21'] = ws.cell(row = r,column = 13).value
    wg['D26'] = ws.cell(row = r,column = 14).value
    wg['F19'] = ws.cell(row = r,column = 15).value
    wg['F20'] = ws.cell(row = r,column = 16).value
    wg['F21'] = ws.cell(row = r,column = 17).value
    wg['F26'] = ws.cell(row = r,column = 18).value
    wg['D27'] = ws.cell(row = r,column = 19).value
    wg['D30'] = ws.cell(row = r,column = 20).value
    wg['D31'] = ws.cell(row = r,column = 21).value
    wg['D32'] = ws.cell(row = r,column = 22).value
    wg['D33'] = ws.cell(row = r,column = 23).value
    wg['D34'] = ws.cell(row = r,column = 24).value
    wg['D35'] = ws.cell(row = r,column = 25).value
    wg['D36'] = ws.cell(row = r,column = 26).value
    Email = ws.cell(row = r,column = 27).value
    print("End of row")
    print(Email)
    Get = (wg['D12'].value)
    wb.save('PAYROLL.xlsx')
    Fi.PDF(str(Get))
    Em.Email(str(Get),Email)
    time.sleep(1)
    r+= 1
    
def Roll(Place, row,column) :
    wg[Place] = ws.cell(row = row,column = column).value